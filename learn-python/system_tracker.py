import time
import os
from pynput import mouse, keyboard
from datetime import datetime
from openpyxl import Workbook, load_workbook
from plyer import notification

# Idle threshold in seconds
IDLE_THRESHOLD = 60
last_active = time.time()
was_idle = False
idle_start = None

# Excel file
excel_file = "idle_log.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Idle Logs"
    ws.append(["Start Time", "End Time", "Idle Duration (seconds)"])
    wb.save(excel_file)

def reset_timer(*args):
    global last_active
    last_active = time.time()

def show_notification(duration):
    notification.notify(
        title="Idle Alert",
        message=f"You were idle for {int(duration)} seconds.",
        timeout=5
    )

def write_to_excel(start, end, duration):
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([
        datetime.fromtimestamp(start).strftime("%Y-%m-%d %H:%M:%S"),
        datetime.fromtimestamp(end).strftime("%Y-%m-%d %H:%M:%S"),
        int(duration)
    ])
    wb.save(excel_file)
    print(f"Idle time logged: {int(duration)} seconds")

def track_idle():
    global last_active, was_idle, idle_start
    print("Tracking idle time. Press Ctrl+C to stop.")
    while True:
        now = time.time()
        idle_time = now - last_active

        if idle_time > IDLE_THRESHOLD and not was_idle:
            idle_start = last_active
            was_idle = True
            show_notification(idle_time)

        elif idle_time < 1 and was_idle:
            idle_end = time.time()
            duration = idle_end - idle_start
            write_to_excel(idle_start, idle_end, duration)
            was_idle = False

        time.sleep(1)

# Start listeners
mouse.Listener(on_move=reset_timer, on_click=reset_timer, on_scroll=reset_timer).start()
keyboard.Listener(on_press=reset_timer).start()

track_idle()
